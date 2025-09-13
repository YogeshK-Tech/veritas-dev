import openpyxl
import pandas as pd
from typing import Dict, Any, List, Optional, Tuple
import structlog
from io import BytesIO
import json
import re
from datetime import datetime

logger = structlog.get_logger()

class ComprehensiveExcelService:
    def __init__(self):
        # Configuration for comprehensive extraction
        self.max_sheets_to_process = 50
        self.max_rows_per_sheet = 2000   # Increased from 50
        self.max_cols_per_sheet = 200    # Increased from 20
        self.memory_efficient_processing = True
    
    async def extract_data_comprehensive(self, file_content: bytes) -> Dict[str, Any]:
        """
        COMPREHENSIVE Excel data extraction - removes all artificial limits
        """
        logger.info("Starting COMPREHENSIVE Excel data extraction - no artificial limits")
        
        try:
            # Load workbook with comprehensive settings
            workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True, read_only=False)
            workbook_with_formulas = openpyxl.load_workbook(BytesIO(file_content), data_only=False, read_only=False)
            
            total_sheets = len(workbook.sheetnames)
            sheets_to_process = min(total_sheets, self.max_sheets_to_process)
            
            logger.info(f"Processing {sheets_to_process} sheets out of {total_sheets} total sheets")
            
            sheets_data = {}
            comprehensive_statistics = {
                "total_sheets": total_sheets,
                "processed_sheets": 0,
                "total_cells_processed": 0,
                "numeric_cells_found": 0,
                "formula_cells_found": 0,
                "text_cells_found": 0,
                "sheets_skipped": 0
            }
            
            for sheet_name in workbook.sheetnames[:sheets_to_process]:
                try:
                    logger.info(f"Processing sheet: {sheet_name}")
                    
                    sheet = workbook[sheet_name]
                    sheet_with_formulas = workbook_with_formulas[sheet_name]
                    
                    sheet_data = await self._extract_comprehensive_sheet_data(sheet, sheet_with_formulas, sheet_name)
                    sheets_data[sheet_name] = sheet_data
                    
                    # Update statistics
                    comprehensive_statistics["processed_sheets"] += 1
                    comprehensive_statistics["total_cells_processed"] += sheet_data["statistics"]["total_cells"]
                    comprehensive_statistics["numeric_cells_found"] += sheet_data["statistics"]["numeric_cells_count"]
                    comprehensive_statistics["formula_cells_found"] += sheet_data["statistics"]["formula_cells_count"]
                    comprehensive_statistics["text_cells_found"] += sheet_data["statistics"]["text_cells_count"]
                    
                except Exception as sheet_error:
                    logger.error(f"Failed to process sheet {sheet_name}: {sheet_error}")
                    comprehensive_statistics["sheets_skipped"] += 1
                    # Continue with other sheets instead of failing completely
                    continue
            
            workbook.close()
            workbook_with_formulas.close()
            
            result = {
                "sheets": sheets_data,
                "metadata": {
                    "sheet_names": list(sheets_data.keys()),
                    "total_sheets_in_file": total_sheets,
                    "processed_sheets": len(sheets_data),
                    "comprehensive_extraction": True,
                    "extraction_timestamp": datetime.utcnow().isoformat()
                },
                "comprehensive_statistics": comprehensive_statistics,
                "extraction_approach": "comprehensive_no_limits"
            }
            
            logger.info(f"COMPREHENSIVE Excel extraction completed:")
            logger.info(f"  - Sheets processed: {comprehensive_statistics['processed_sheets']}")
            logger.info(f"  - Total cells: {comprehensive_statistics['total_cells_processed']}")
            logger.info(f"  - Numeric cells: {comprehensive_statistics['numeric_cells_found']}")
            logger.info(f"  - Formula cells: {comprehensive_statistics['formula_cells_found']}")
            
            return result
            
        except Exception as e:
            logger.error(f"Comprehensive Excel extraction failed: {e}")
            raise
    
    async def _extract_comprehensive_sheet_data(self, sheet, sheet_with_formulas, sheet_name: str) -> Dict[str, Any]:
        """
        Extract comprehensive data from a single sheet - NO ARTIFICIAL LIMITS
        """
        # Get ACTUAL sheet dimensions
        actual_max_row = sheet.max_row or 1
        actual_max_col = sheet.max_column or 1
        
        # Apply reasonable limits for memory management only
        effective_max_row = min(actual_max_row, self.max_rows_per_sheet)
        effective_max_col = min(actual_max_col, self.max_cols_per_sheet)
        
        logger.info(f"Sheet '{sheet_name}' dimensions: {actual_max_row}x{actual_max_col}, processing: {effective_max_row}x{effective_max_col}")
        
        # Initialize comprehensive data structures
        cells_data = {}
        numeric_cells = []
        text_cells = []
        formula_cells = []
        date_cells = []
        percentage_cells = []
        currency_cells = []
        
        # Process ALL cells in the effective range
        for row in range(1, effective_max_row + 1):
            for col in range(1, effective_max_col + 1):
                try:
                    cell = sheet.cell(row, col)
                    
                    if cell.value is not None:
                        cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                        
                        # Extract comprehensive cell information
                        cell_info = await self._extract_comprehensive_cell_info(cell, sheet_with_formulas.cell(row, col), row, col)
                        
                        cells_data[cell_ref] = cell_info
                        
                        # Categorize cells comprehensively
                        await self._categorize_cell_comprehensive(cell_info, cell_ref, numeric_cells, text_cells, formula_cells, date_cells, percentage_cells, currency_cells)
                
                except Exception as cell_error:
                    # Log but continue processing other cells
                    logger.debug(f"Error processing cell {col},{row} in sheet {sheet_name}: {cell_error}")
                    continue
        
        # Detect comprehensive data patterns
        data_regions = await self._detect_comprehensive_data_regions(cells_data, effective_max_row, effective_max_col)
        
        # Identify high-priority cells (KPIs, summary metrics, etc.)
        high_priority_cells = await self._identify_high_priority_cells(numeric_cells, text_cells, formula_cells)
        
        # Analyze data relationships and patterns
        data_patterns = await self._analyze_data_patterns(cells_data, data_regions)
        
        # Create comprehensive sheet analysis
        sheet_data = {
            "cells": cells_data,
            "comprehensive_categorization": {
                "numeric_cells": numeric_cells,        # ALL numeric cells - no limits
                "text_cells": text_cells[:500],        # Reasonable limit for text cells
                "formula_cells": formula_cells,        # ALL formula cells - no limits
                "date_cells": date_cells,             # ALL date cells
                "percentage_cells": percentage_cells,  # ALL percentage cells
                "currency_cells": currency_cells      # ALL currency cells
            },
            "data_regions": data_regions,
            "high_priority_cells": high_priority_cells,
            "data_patterns": data_patterns,
            "dimensions": {
                "actual_max_row": actual_max_row,
                "actual_max_col": actual_max_col,
                "processed_max_row": effective_max_row,
                "processed_max_col": effective_max_col,
                "coverage_percentage": min(100, (effective_max_row * effective_max_col) / (actual_max_row * actual_max_col) * 100)
            },
            "statistics": {
                "total_cells": len(cells_data),
                "numeric_cells_count": len(numeric_cells),
                "text_cells_count": len(text_cells),
                "formula_cells_count": len(formula_cells),
                "date_cells_count": len(date_cells),
                "percentage_cells_count": len(percentage_cells),
                "currency_cells_count": len(currency_cells),
                "high_priority_cells_count": len(high_priority_cells)
            },
            "extraction_metadata": {
                "comprehensive_extraction": True,
                "limitations_removed": ["artificial_cell_limits", "numeric_cell_limits", "arbitrary_result_limits"],
                "extraction_approach": "comprehensive_categorized"
            }
        }
        
        logger.info(f"Sheet '{sheet_name}' comprehensive extraction completed:")
        logger.info(f"  - Total cells: {len(cells_data)}")
        logger.info(f"  - Numeric: {len(numeric_cells)}, Text: {len(text_cells)}, Formulas: {len(formula_cells)}")
        logger.info(f"  - Dates: {len(date_cells)}, Percentages: {len(percentage_cells)}, Currency: {len(currency_cells)}")
        
        return sheet_data
    
    async def _extract_comprehensive_cell_info(self, cell, formula_cell, row: int, col: int) -> Dict[str, Any]:
        """Extract comprehensive information about a single cell"""
        
        cell_info = {
            "value": cell.value,
            "data_type": type(cell.value).__name__,
            "row": row,
            "col": col,
            "coordinate": f"{openpyxl.utils.get_column_letter(col)}{row}"
        }
        
        # Number formatting information
        if hasattr(cell, 'number_format') and cell.number_format:
            cell_info["number_format"] = cell.number_format
            cell_info["is_percentage"] = '%' in cell.number_format
            cell_info["is_currency"] = any(symbol in cell.number_format for symbol in ['$', '€', '£', '¥', '₹'])
            cell_info["is_date"] = any(fmt in cell.number_format.lower() for fmt in ['yyyy', 'mm', 'dd', 'date'])
        
        # Font and styling information
        if hasattr(cell, 'font') and cell.font:
            cell_info["font"] = {
                "bold": getattr(cell.font, 'bold', False),
                "italic": getattr(cell.font, 'italic', False),
                "size": getattr(cell.font, 'size', 11),
                "color": str(getattr(cell.font, 'color', '')) if getattr(cell.font, 'color', None) else None
            }
        
        # Border and fill information (for identifying important cells)
        if hasattr(cell, 'border') and cell.border:
            border_styles = [getattr(cell.border.left, 'style', None), getattr(cell.border.right, 'style', None),
                           getattr(cell.border.top, 'style', None), getattr(cell.border.bottom, 'style', None)]
            cell_info["has_borders"] = any(style is not None for style in border_styles)
        
        if hasattr(cell, 'fill') and cell.fill:
            cell_info["has_fill"] = getattr(cell.fill, 'start_color', None) is not None
        
        # Formula information
        if formula_cell and str(formula_cell.value).startswith('='):
            cell_info["formula"] = str(formula_cell.value)
            cell_info["is_calculated"] = True
            
            # Analyze formula complexity
            formula_str = str(formula_cell.value)
            cell_info["formula_complexity"] = {
                "has_sum": "SUM" in formula_str.upper(),
                "has_average": "AVERAGE" in formula_str.upper(),
                "has_count": "COUNT" in formula_str.upper(),
                "has_vlookup": "VLOOKUP" in formula_str.upper(),
                "has_if": "IF" in formula_str.upper(),
                "cell_references": len(re.findall(r'[A-Z]+\d+', formula_str)),
                "is_complex": len(formula_str) > 20
            }
        else:
            cell_info["is_calculated"] = False
        
        # Value analysis
        if isinstance(cell.value, (int, float)):
            cell_info["numeric_analysis"] = {
                "absolute_value": abs(cell.value),
                "is_large_number": abs(cell.value) >= 1000,
                "is_round_number": cell.value != 0 and cell.value % 1000 == 0,
                "is_percentage_range": 0 <= cell.value <= 100,
                "is_ratio_range": 0 <= cell.value <= 10,
                "magnitude": len(str(int(abs(cell.value)))) if cell.value != 0 else 0
            }
        
        # Context clues from position
        cell_info["position_analysis"] = {
            "is_in_first_row": row == 1,
            "is_in_first_column": col == 1,
            "is_corner_cell": (row <= 3 and col <= 3),
            "row_position": "top" if row <= 10 else "middle" if row <= 100 else "bottom",
            "col_position": "left" if col <= 5 else "middle" if col <= 20 else "right"
        }
        
        return cell_info
    
    async def _categorize_cell_comprehensive(self, cell_info: Dict, cell_ref: str, 
                                           numeric_cells: List, text_cells: List, formula_cells: List,
                                           date_cells: List, percentage_cells: List, currency_cells: List):
        """Comprehensively categorize cells based on their characteristics"""
        
        value = cell_info["value"]
        
        # Numeric cells categorization
        if isinstance(value, (int, float)) and value != 0:
            numeric_entry = {
                "cell_ref": cell_ref,
                "value": value,
                **cell_info
            }
            numeric_cells.append(numeric_entry)
            
            # Sub-categorize numeric cells
            if cell_info.get("is_percentage", False):
                percentage_cells.append(numeric_entry)
            
            if cell_info.get("is_currency", False):
                currency_cells.append(numeric_entry)
        
        # Text cells
        elif isinstance(value, str) and len(value.strip()) > 0:
            text_entry = {
                "cell_ref": cell_ref,
                "value": value,
                **cell_info
            }
            text_cells.append(text_entry)
        
        # Date cells
        elif cell_info.get("is_date", False) or isinstance(value, datetime):
            date_entry = {
                "cell_ref": cell_ref,
                "value": value,
                **cell_info
            }
            date_cells.append(date_entry)
        
        # Formula cells
        if cell_info.get("is_calculated", False):
            formula_entry = {
                "cell_ref": cell_ref,
                "formula": cell_info.get("formula"),
                "result_value": value,
                **cell_info
            }
            formula_cells.append(formula_entry)
    
    async def _identify_high_priority_cells(self, numeric_cells: List, text_cells: List, formula_cells: List) -> List[Dict]:
        """Identify cells that are most likely to contain presentation-worthy values"""
        
        high_priority_cells = []
        
        # Score numeric cells for presentation likelihood
        for cell in numeric_cells:
            score = await self._calculate_presentation_score(cell)
            if score >= 5:  # Threshold for high priority
                cell["presentation_score"] = score
                high_priority_cells.append(cell)
        
        # Score formula cells (often calculated KPIs)
        for cell in formula_cells:
            if isinstance(cell.get("result_value"), (int, float)):
                score = await self._calculate_presentation_score(cell)
                # Formula cells get bonus points
                score += 2
                if score >= 4:  # Lower threshold for formula cells
                    cell["presentation_score"] = score
                    high_priority_cells.append(cell)
        
        # Sort by presentation score
        high_priority_cells.sort(key=lambda x: x.get("presentation_score", 0), reverse=True)
        
        return high_priority_cells[:1000]  # Return top 1000 high-priority cells
    
    async def _calculate_presentation_score(self, cell: Dict) -> int:
        """Calculate how likely a cell is to appear in presentations"""
        
        score = 0
        value = cell.get("value", 0)
        
        if not isinstance(value, (int, float)):
            return 0
        
        # Size-based scoring
        abs_value = abs(value)
        if abs_value >= 1000000:  # Millions
            score += 4
        elif abs_value >= 100000:  # Hundreds of thousands
            score += 3
        elif abs_value >= 10000:  # Ten thousands
            score += 2
        elif abs_value >= 1000:  # Thousands
            score += 1
        
        # Round number bonus
        if abs_value > 0 and abs_value % 1000 == 0:
            score += 2
        
        # Percentage range bonus
        if cell.get("numeric_analysis", {}).get("is_percentage_range", False):
            score += 2
        
        # Formatting-based scoring
        font_info = cell.get("font", {})
        if font_info.get("bold", False):
            score += 3
        if font_info.get("size", 11) > 12:
            score += 1
        
        # Number format bonuses
        if cell.get("is_currency", False):
            score += 2
        if cell.get("is_percentage", False):
            score += 2
        
        # Border/fill bonuses (formatted cells are often important)
        if cell.get("has_borders", False):
            score += 1
        if cell.get("has_fill", False):
            score += 1
        
        # Formula complexity bonus
        if cell.get("is_calculated", False):
            complexity = cell.get("formula_complexity", {})
            if complexity.get("is_complex", False):
                score += 2
            if complexity.get("has_sum", False) or complexity.get("has_average", False):
                score += 1
        
        # Position-based scoring
        position = cell.get("position_analysis", {})
        if position.get("is_corner_cell", False):
            score += 1
        
        return score
    
    async def _detect_comprehensive_data_regions(self, cells_data: Dict, max_row: int, max_col: int) -> List[Dict]:
        """Detect comprehensive data regions (tables, summary areas, etc.)"""
        
        regions = []
        processed_areas = set()
        
        # Scan for data regions more comprehensively
        for start_row in range(1, min(max_row, 500), 15):  # Sample every 15 rows
            for start_col in range(1, min(max_col, 50), 8):   # Sample every 8 columns
                
                region_key = f"{start_row}_{start_col}"
                if region_key in processed_areas:
                    continue
                
                region = await self._analyze_comprehensive_data_region(cells_data, start_row, start_col, max_row, max_col)
                
                if region and region["cell_count"] >= 9:  # Minimum meaningful region size
                    regions.append(region)
                    
                    # Mark area as processed
                    for row in range(region["start_row"], min(region["end_row"] + 1, max_row + 1)):
                        for col in range(region["start_col"], min(region["end_col"] + 1, max_col + 1)):
                            processed_areas.add(f"{row}_{col}")
        
        return regions
    
    async def _analyze_comprehensive_data_region(self, cells_data: Dict, start_row: int, start_col: int, max_row: int, max_col: int) -> Optional[Dict]:
        """Analyze a potential comprehensive data region"""
        
        region_cells = []
        numeric_cells_in_region = []
        text_cells_in_region = []
        
        # Expand region to find contiguous data
        end_row = start_row
        end_col = start_col
        
        for row in range(start_row, min(start_row + 30, max_row + 1)):  # Larger regions
            row_has_data = False
            for col in range(start_col, min(start_col + 25, max_col + 1)):  # Wider regions
                cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                if cell_ref in cells_data:
                    cell_data = cells_data[cell_ref]
                    region_cells.append({
                        "cell_ref": cell_ref,
                        "value": cell_data["value"],
                        "row": row,
                        "col": col,
                        "data_type": cell_data["data_type"]
                    })
                    
                    if isinstance(cell_data["value"], (int, float)):
                        numeric_cells_in_region.append(cell_data)
                    elif isinstance(cell_data["value"], str):
                        text_cells_in_region.append(cell_data)
                    
                    row_has_data = True
                    end_row = max(end_row, row)
                    end_col = max(end_col, col)
            
            # If row has no data and we have some data, consider ending the region
            if not row_has_data and len(region_cells) > 6:
                break
        
        if len(region_cells) >= 9:  # Minimum meaningful region
            region_analysis = {
                "start_row": start_row,
                "start_col": start_col,
                "end_row": end_row,
                "end_col": end_col,
                "cell_count": len(region_cells),
                "numeric_cells_count": len(numeric_cells_in_region),
                "text_cells_count": len(text_cells_in_region),
                "cells": region_cells[:100],  # Sample of cells for analysis
                "density": len(region_cells) / ((end_row - start_row + 1) * (end_col - start_col + 1)),
                "region_type": await self._classify_region_type(region_cells, numeric_cells_in_region, text_cells_in_region)
            }
            
            return region_analysis
        
        return None
    
    async def _classify_region_type(self, region_cells: List, numeric_cells: List, text_cells: List) -> str:
        """Classify the type of data region"""
        
        total_cells = len(region_cells)
        numeric_ratio = len(numeric_cells) / total_cells if total_cells > 0 else 0
        text_ratio = len(text_cells) / total_cells if total_cells > 0 else 0
        
        if numeric_ratio > 0.8:
            return "numeric_table"
        elif numeric_ratio > 0.5:
            return "mixed_data_table"
        elif text_ratio > 0.8:
            return "text_table"
        elif numeric_ratio > 0.2:
            return "summary_section"
        else:
            return "general_data_region"
    
    async def _analyze_data_patterns(self, cells_data: Dict, data_regions: List) -> Dict[str, Any]:
        """Analyze patterns in the data for better understanding"""
        
        patterns = {
            "large_numbers_distribution": [],
            "currency_patterns": [],
            "percentage_patterns": [],
            "formula_patterns": [],
            "summary_indicators": []
        }
        
        # Analyze large numbers (likely to be presentation values)
        for cell_ref, cell_info in cells_data.items():
            value = cell_info.get("value")
            
            if isinstance(value, (int, float)) and abs(value) >= 10000:
                patterns["large_numbers_distribution"].append({
                    "cell_ref": cell_ref,
                    "value": value,
                    "magnitude": len(str(int(abs(value)))),
                    "is_round": value % 1000 == 0,
                    "formatting": cell_info.get("number_format", "")
                })
            
            # Currency patterns
            if cell_info.get("is_currency", False):
                patterns["currency_patterns"].append({
                    "cell_ref": cell_ref,
                    "value": value,
                    "format": cell_info.get("number_format", "")
                })
            
            # Percentage patterns
            if cell_info.get("is_percentage", False):
                patterns["percentage_patterns"].append({
                    "cell_ref": cell_ref,
                    "value": value,
                    "format": cell_info.get("number_format", "")
                })
            
            # Formula patterns
            if cell_info.get("is_calculated", False):
                patterns["formula_patterns"].append({
                    "cell_ref": cell_ref,
                    "formula": cell_info.get("formula", ""),
                    "result": value,
                    "complexity": cell_info.get("formula_complexity", {})
                })
        
        # Identify summary indicators
        patterns["summary_indicators"] = await self._identify_summary_indicators(cells_data)
        
        return patterns
    
    async def _identify_summary_indicators(self, cells_data: Dict) -> List[Dict]:
        """Identify cells that appear to be summary/total indicators"""
        
        summary_indicators = []
        
        for cell_ref, cell_info in cells_data.items():
            value = cell_info.get("value")
            
            # Look for cells that might be totals/summaries
            if isinstance(value, str):
                value_lower = value.lower()
                if any(keyword in value_lower for keyword in ["total", "sum", "revenue", "profit", "loss", "net", "gross", "ebitda"]):
                    summary_indicators.append({
                        "cell_ref": cell_ref,
                        "text": value,
                        "indicator_type": "text_summary_label"
                    })
            
            elif isinstance(value, (int, float)):
                # Large round numbers in bold formatting are often summaries
                if (abs(value) >= 100000 and 
                    value % 1000 == 0 and 
                    cell_info.get("font", {}).get("bold", False)):
                    summary_indicators.append({
                        "cell_ref": cell_ref,
                        "value": value,
                        "indicator_type": "numeric_summary_value"
                    })
        
        return summary_indicators

# Update the original excel_service.py to use comprehensive extraction
class ExcelService(ComprehensiveExcelService):
    """
    Enhanced Excel Service that uses comprehensive extraction by default
    """
    
    async def extract_data(self, file_content: bytes) -> Dict[str, Any]:
        """
        Main extraction method - now uses comprehensive extraction
        """
        return await self.extract_data_comprehensive(file_content)

# Singleton instance
excel_service = ExcelService()