import google.generativeai as genai
from typing import Dict, Any, List, Tuple, Optional
import json
import base64
import structlog
from PIL import Image
import io
import time
import asyncio
import os
import fitz  # PyMuPDF
from decouple import config
import re
from datetime import datetime
import math

# Configure logging
logger = structlog.get_logger()

class EnhancedGeminiService:
    def __init__(self):
        # Get API key from environment
        api_key = config('GOOGLE_API_KEY', default=None)
        
        if not api_key:
            raise ValueError("GOOGLE_API_KEY is required for Gemini 2.5 Pro")
        
        genai.configure(api_key=api_key)
        self.model = genai.GenerativeModel('gemini-2.0-flash-exp')
        self.ai_enabled = True
        
        # Configuration for comprehensive extraction
        self.max_cells_per_batch = 200  # Increased from previous limits
        self.max_sheets_per_workbook = 50  # Process up to 50 sheets
        self.max_rows_per_sheet = 1000  # Up from 50
        self.max_cols_per_sheet = 100   # Up from 20
        
        logger.info("Enhanced Gemini 2.5 Pro Service initialized with comprehensive extraction settings")

    async def extract_comprehensive_pdf_data(self, pdf_path: str) -> Dict[str, Any]:
        """
        Comprehensive PDF extraction using Gemini 2.5 Pro with bounding boxes
        """
        logger.info(f"Starting comprehensive PDF extraction: {pdf_path}")
        
        try:
            doc = fitz.open(pdf_path)
            page_analyses = []
            
            for page_num in range(len(doc)):
                page = doc[page_num]
                
                # Extract high-quality page image with coordinates
                page_data = await self._extract_page_with_coordinates(page, page_num + 1)
                page_analyses.append(page_data)
                
                # Rate limiting for Gemini API
                await asyncio.sleep(1.5)
            
            doc.close()
            
            # Synthesize complete document analysis
            comprehensive_data = await self._synthesize_document_analysis(page_analyses)
            
            logger.info(f"PDF extraction completed: {len(comprehensive_data.get('all_extracted_values', []))} values found")
            return comprehensive_data
            
        except Exception as e:
            logger.error(f"PDF extraction failed: {e}")
            raise

    async def _extract_page_with_coordinates(self, page, page_num: int) -> Dict[str, Any]:
        """
        Extract page data with precise coordinate mapping using Gemini 2.5 Pro
        """
        try:
            # Convert page to high-resolution image
            mat = fitz.Matrix(2.0, 2.0)  # High quality for better extraction
            pix = page.get_pixmap(matrix=mat)
            img_data = pix.tobytes("png")
            
            # Convert to PIL Image for Gemini
            image = Image.open(io.BytesIO(img_data))
            
            # Enhanced prompt for better extraction
            prompt = f"""
Analyze this presentation slide (page {page_num}) and extract ALL numerical values, financial metrics, percentages, dates, and quantitative data.

CRITICAL: Extract EVERY number you can find on this slide, including:
- Revenue figures, costs, profits
- Percentages, ratios, growth rates  
- Dates, years, quarters
- Counts, quantities, volumes
- Currency amounts in any format
- Statistical data, KPIs, metrics

For each number found, provide:
- Exact value as displayed
- Business context explaining what it represents
- Normalized coordinates [x1, y1, x2, y2] on 0-1 scale
- Data type classification

Return ONLY valid JSON in this exact format:
{{
    "page_number": {page_num},
    "page_dimensions": {{"width": {image.width}, "height": {image.height}}},
    "extracted_values": [
        {{
            "id": "value_{page_num}_001",
            "value": "exact_number_as_displayed",
            "normalized_value": "cleaned_numeric_format",
            "data_type": "currency|percentage|count|ratio|date|metric",
            "coordinates": {{
                "bounding_box": [0.1, 0.2, 0.3, 0.4],
                "confidence": 0.9
            }},
            "business_context": {{
                "semantic_meaning": "detailed_description_of_what_this_number_represents",
                "business_category": "revenue|costs|growth|operational|financial|market",
                "presentation_priority": "primary|secondary|supporting",
                "calculation_type": "absolute|percentage|ratio|growth_rate|other"
            }},
            "confidence": 0.9
        }}
    ]
}}

IMPORTANT: Be thorough - extract ALL visible numbers, not just the prominent ones. Return only valid JSON.
"""

            response = self.model.generate_content([prompt, image])
            result = await self._parse_gemini_json_response_robust(response.text, f"page_{page_num}")
            
            # Validate and enhance coordinates
            result = self._validate_and_enhance_coordinates(result, image.size)
            
            logger.info(f"Page {page_num}: Extracted {len(result.get('extracted_values', []))} values")
            return result
            
        except Exception as e:
            logger.error(f"Page {page_num} extraction failed: {e}")
            return {
                "page_number": page_num,
                "page_dimensions": {"width": 800, "height": 600},
                "extracted_values": [],
                "error": str(e)
            }

    async def analyze_excel_comprehensive(self, excel_path: str) -> Dict[str, Any]:
        """
        COMPLETELY REDESIGNED comprehensive Excel analysis - no artificial limits
        """
        logger.info(f"Starting COMPREHENSIVE Excel analysis (no limits): {excel_path}")
        
        try:
            import openpyxl
            
            # Load workbook with both data and formulas
            wb_data = openpyxl.load_workbook(excel_path, data_only=True)
            wb_formulas = openpyxl.load_workbook(excel_path, data_only=False)
            
            total_sheets = len(wb_data.sheetnames)
            sheets_to_process = min(total_sheets, self.max_sheets_per_workbook)
            
            logger.info(f"Processing {sheets_to_process} sheets out of {total_sheets} total sheets")
            
            # Process ALL sheets (up to reasonable limit)
            sheet_analyses = []
            all_potential_sources = []
            
            for i, sheet_name in enumerate(wb_data.sheetnames[:sheets_to_process]):
                logger.info(f"Processing Excel sheet {i+1}/{sheets_to_process}: {sheet_name}")
                
                sheet_data = wb_data[sheet_name]
                sheet_formulas = wb_formulas[sheet_name]
                
                # Extract FULL sheet structure (no artificial limits)
                sheet_structure = await self._extract_full_excel_sheet_structure(sheet_data, sheet_formulas)
                
                # Process in batches to handle large sheets
                sheet_analysis = await self._analyze_excel_sheet_comprehensive_batched(sheet_name, sheet_structure)
                sheet_analyses.append(sheet_analysis)
                
                # Collect all potential sources
                sources = sheet_analysis.get("potential_sources", [])
                for source in sources:
                    source["source_sheet"] = sheet_name
                    source["sheet_index"] = i
                all_potential_sources.extend(sources)
                
                # Rate limiting between sheets
                await asyncio.sleep(1)
            
            # Synthesize workbook analysis with ALL data
            workbook_analysis = await self._synthesize_comprehensive_excel_workbook(sheet_analyses, all_potential_sources)
            
            total_sources = len(all_potential_sources)
            logger.info(f"COMPREHENSIVE Excel analysis completed: {total_sources} potential sources identified across {sheets_to_process} sheets")
            
            return workbook_analysis
            
        except Exception as e:
            logger.error(f"Comprehensive Excel analysis failed: {e}")
            raise

    async def _extract_full_excel_sheet_structure(self, sheet_data, sheet_formulas) -> Dict[str, Any]:
        """Extract FULL Excel sheet structure - no artificial row/column limits"""
        import openpyxl
        
        # Get actual sheet dimensions
        actual_max_row = sheet_data.max_row or 1
        actual_max_col = sheet_data.max_column or 1
        
        # Apply reasonable limits for memory management
        max_row = min(actual_max_row, self.max_rows_per_sheet)
        max_col = min(actual_max_col, self.max_cols_per_sheet)
        
        logger.info(f"Sheet dimensions: {actual_max_row}x{actual_max_col}, processing: {max_row}x{max_col}")
        
        cells_data = {}
        numeric_cells = []
        text_cells = []
        formula_cells = []
        
        # Extract ALL relevant cells
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                try:
                    cell_data = sheet_data.cell(row, col)
                    
                    if cell_data.value is not None:
                        cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                        
                        cell_info = {
                            "value": cell_data.value,
                            "data_type": type(cell_data.value).__name__,
                            "row": row,
                            "col": col,
                            "number_format": getattr(cell_data, 'number_format', None),
                            "font_bold": getattr(cell_data.font, 'bold', False) if cell_data.font else False,
                            "font_size": getattr(cell_data.font, 'size', 11) if cell_data.font else 11
                        }
                        
                        # Add formula if exists
                        try:
                            cell_formula = sheet_formulas.cell(row, col)
                            if cell_formula.value and str(cell_formula.value).startswith('='):
                                cell_info["formula"] = str(cell_formula.value)
                                formula_cells.append({
                                    "cell_ref": cell_ref,
                                    "formula": cell_info["formula"],
                                    "result_value": cell_data.value,
                                    **cell_info
                                })
                        except:
                            pass
                        
                        cells_data[cell_ref] = cell_info
                        
                        # Categorize cells by type
                        if isinstance(cell_data.value, (int, float)) and abs(cell_data.value) > 0:
                            numeric_cells.append({
                                "cell_ref": cell_ref,
                                "value": cell_data.value,
                                **cell_info
                            })
                        elif isinstance(cell_data.value, str) and len(cell_data.value.strip()) > 0:
                            text_cells.append({
                                "cell_ref": cell_ref,
                                "value": cell_data.value,
                                **cell_info
                            })
                
                except Exception as e:
                    # Skip problematic cells but continue processing
                    continue
        
        # Detect data regions and important patterns
        data_regions = self._detect_comprehensive_data_regions(cells_data, max_row, max_col)
        
        # Identify high-value cells (likely to be KPIs)
        high_value_cells = self._identify_high_value_cells(numeric_cells, text_cells)
        
        structure = {
            "cells": cells_data,
            "numeric_cells": numeric_cells,  # NO LIMITS - include ALL numeric cells
            "text_cells": text_cells[:100],  # Limit text cells for performance
            "formula_cells": formula_cells,
            "high_value_cells": high_value_cells,
            "data_regions": data_regions,
            "dimensions": {
                "actual_max_row": actual_max_row,
                "actual_max_col": actual_max_col,
                "processed_max_row": max_row,
                "processed_max_col": max_col
            },
            "statistics": {
                "total_cells": len(cells_data),
                "numeric_cells_count": len(numeric_cells),
                "text_cells_count": len(text_cells),
                "formula_cells_count": len(formula_cells)
            }
        }
        
        logger.info(f"Extracted {len(numeric_cells)} numeric cells, {len(text_cells)} text cells, {len(formula_cells)} formula cells")
        
        return structure

    async def _analyze_excel_sheet_comprehensive_batched(self, sheet_name: str, sheet_structure: Dict) -> Dict[str, Any]:
        """Analyze Excel sheet using intelligent batching for comprehensive coverage"""
        
        numeric_cells = sheet_structure["numeric_cells"]
        high_value_cells = sheet_structure.get("high_value_cells", [])
        
        logger.info(f"Analyzing sheet '{sheet_name}' with {len(numeric_cells)} numeric cells using intelligent batching")
        
        # Process high-value cells first (most likely to be presentation values)
        all_potential_sources = []
        
        if high_value_cells:
            high_value_sources = await self._analyze_cell_batch_with_gemini(
                sheet_name, high_value_cells, "high_value", batch_index=0
            )
            all_potential_sources.extend(high_value_sources)
        
        # Process remaining numeric cells in batches
        remaining_cells = [cell for cell in numeric_cells if cell not in high_value_cells]
        
        batch_size = self.max_cells_per_batch
        total_batches = math.ceil(len(remaining_cells) / batch_size)
        
        for batch_index in range(total_batches):
            start_idx = batch_index * batch_size
            end_idx = min(start_idx + batch_size, len(remaining_cells))
            batch_cells = remaining_cells[start_idx:end_idx]
            
            logger.info(f"Processing batch {batch_index + 1}/{total_batches} for sheet '{sheet_name}' ({len(batch_cells)} cells)")
            
            batch_sources = await self._analyze_cell_batch_with_gemini(
                sheet_name, batch_cells, "standard", batch_index + 1
            )
            all_potential_sources.extend(batch_sources)
            
            # Rate limiting between batches
            await asyncio.sleep(0.5)
        
        # Sort by presentation likelihood
        all_potential_sources.sort(key=lambda x: x.get("presentation_likelihood", 0), reverse=True)
        
        analysis_result = {
            "sheet_name": sheet_name,
            "potential_sources": all_potential_sources,  # NO LIMITS - include ALL sources
            "analysis_metadata": {
                "total_numeric_cells_analyzed": len(numeric_cells),
                "total_batches_processed": total_batches + (1 if high_value_cells else 0),
                "high_value_cells_count": len(high_value_cells),
                "comprehensive_coverage": True
            }
        }
        
        logger.info(f"Sheet '{sheet_name}' analysis completed: {len(all_potential_sources)} potential sources identified")
        
        return analysis_result

    async def _analyze_cell_batch_with_gemini(self, sheet_name: str, cells_batch: List[Dict], batch_type: str, batch_index: int) -> List[Dict]:
        """Analyze a batch of cells with Gemini for comprehensive extraction"""
        
        if not cells_batch:
            return []
        
        try:
            # Prepare batch data for analysis
            batch_data = []
            for cell in cells_batch:
                batch_data.append({
                    "cell_ref": cell["cell_ref"],
                    "value": cell["value"],
                    "data_type": cell["data_type"],
                    "formula": cell.get("formula"),
                    "font_bold": cell.get("font_bold", False),
                    "number_format": cell.get("number_format")
                })
            
            prompt = f"""
Analyze these Excel cells from sheet '{sheet_name}' (batch {batch_index}, type: {batch_type}) to identify values likely to appear in business presentations.

CELLS TO ANALYZE:
{json.dumps(batch_data, indent=1)}

For each cell that could potentially appear in a presentation, determine:
1. How likely it is to be referenced in presentations (0.0 to 1.0)
2. What business context it represents
3. What type of data it is

Focus on:
- Financial metrics (revenue, costs, profits, margins)
- Growth rates and percentages
- Key performance indicators
- Market data and statistics
- Operational metrics
- Strategic numbers

Return ONLY valid JSON:
{{
    "batch_analysis": [
        {{
            "cell_reference": "A1",
            "value": "actual_value",
            "business_context": "detailed_business_meaning",
            "presentation_likelihood": 0.9,
            "data_type": "currency|percentage|count|ratio|metric",
            "value_category": "revenue|cost|growth|operational|market|strategic",
            "reasoning": "why_this_might_appear_in_presentations"
        }}
    ]
}}

Return only valid JSON with comprehensive analysis.
"""

            response = self.model.generate_content(prompt)
            result = await self._parse_gemini_json_response_robust(response.text, f"excel_batch_{sheet_name}_{batch_index}")
            
            batch_analysis = result.get('batch_analysis', [])
            
            # Filter for likely presentation values (threshold can be adjusted)
            presentation_sources = [
                source for source in batch_analysis 
                if source.get("presentation_likelihood", 0) >= 0.3  # Lower threshold for comprehensive coverage
            ]
            
            logger.info(f"Batch {batch_index} ({batch_type}): {len(presentation_sources)} presentation-worthy sources from {len(cells_batch)} cells")
            
            return presentation_sources
            
        except Exception as e:
            logger.error(f"Batch analysis failed for {sheet_name} batch {batch_index}: {e}")
            # Return basic structure for failed batches
            return [
                {
                    "cell_reference": cell["cell_ref"],
                    "value": cell["value"],
                    "business_context": "Analysis failed - manual review needed",
                    "presentation_likelihood": 0.5,
                    "data_type": "unknown",
                    "error": str(e)
                }
                for cell in cells_batch[:10]  # Return first 10 cells as fallback
            ]

    def _identify_high_value_cells(self, numeric_cells: List[Dict], text_cells: List[Dict]) -> List[Dict]:
        """Identify cells most likely to contain presentation-worthy values"""
        
        high_value_cells = []
        
        for cell in numeric_cells:
            score = 0
            
            # Large numbers often indicate important metrics
            abs_value = abs(cell["value"])
            if abs_value >= 1000000:  # Millions
                score += 3
            elif abs_value >= 100000:  # Hundreds of thousands
                score += 2
            elif abs_value >= 10000:  # Ten thousands
                score += 1
            
            # Round numbers often indicate calculated/summary metrics
            if abs_value > 0 and abs_value % 1000 == 0:
                score += 1
            
            # Percentages between 0-100 (for growth rates, margins, etc.)
            if 0 < abs_value <= 100 and cell.get("number_format", "").find("%") != -1:
                score += 2
            
            # Bold formatting often indicates important values
            if cell.get("font_bold", False):
                score += 2
            
            # Larger font sizes indicate importance
            font_size = cell.get("font_size", 11)
            if font_size > 12:
                score += 1
            
            # Currency formatting indicates financial metrics
            number_format = cell.get("number_format", "")
            if any(currency_symbol in number_format for currency_symbol in ["$", "€", "£", "¥"]):
                score += 2
            
            # Formula cells might be calculated KPIs
            if cell.get("formula"):
                score += 1
            
            # If score is high enough, consider it high-value
            if score >= 3:
                cell["importance_score"] = score
                high_value_cells.append(cell)
        
        # Sort by importance score
        high_value_cells.sort(key=lambda x: x.get("importance_score", 0), reverse=True)
        
        # Return top candidates (but still allow many)
        return high_value_cells[:500]  # Much higher limit than before

    def _detect_comprehensive_data_regions(self, cells_data: Dict, max_row: int, max_col: int) -> List[Dict]:
        """Detect data regions in the sheet for better context understanding"""
        
        regions = []
        processed_cells = set()
        
        # Look for rectangular data regions
        for start_row in range(1, min(max_row, 200), 10):  # Sample every 10 rows
            for start_col in range(1, min(max_col, 50), 5):   # Sample every 5 columns
                
                if f"{start_col}_{start_row}" in processed_cells:
                    continue
                
                region = self._analyze_data_region(cells_data, start_row, start_col, max_row, max_col)
                
                if region and region["cell_count"] >= 6:  # Minimum table size
                    regions.append(region)
                    
                    # Mark cells as processed
                    for row in range(region["start_row"], region["end_row"] + 1):
                        for col in range(region["start_col"], region["end_col"] + 1):
                            processed_cells.add(f"{col}_{row}")
        
        return regions

    def _analyze_data_region(self, cells_data: Dict, start_row: int, start_col: int, max_row: int, max_col: int) -> Dict[str, Any]:
        """Analyze a potential data region"""
        import openpyxl
        
        region_cells = []
        end_row = start_row
        end_col = start_col
        
        # Expand region to find contiguous data
        for row in range(start_row, min(start_row + 20, max_row + 1)):
            row_has_data = False
            for col in range(start_col, min(start_col + 15, max_col + 1)):
                cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                if cell_ref in cells_data:
                    region_cells.append({
                        "cell_ref": cell_ref,
                        "value": cells_data[cell_ref]["value"],
                        "row": row,
                        "col": col
                    })
                    row_has_data = True
                    end_row = max(end_row, row)
                    end_col = max(end_col, col)
            
            # If row has no data, stop expanding
            if not row_has_data and len(region_cells) > 0:
                break
        
        if len(region_cells) >= 6:
            return {
                "start_row": start_row,
                "start_col": start_col,
                "end_row": end_row,
                "end_col": end_col,
                "cell_count": len(region_cells),
                "cells": region_cells[:50],  # Sample of cells for context
                "density": len(region_cells) / ((end_row - start_row + 1) * (end_col - start_col + 1))
            }
        
        return None

    async def _synthesize_comprehensive_excel_workbook(self, sheet_analyses: List[Dict], all_potential_sources: List[Dict]) -> Dict[str, Any]:
        """Synthesize comprehensive workbook analysis with ALL extracted data"""
        
        try:
            # Sort all sources by presentation likelihood
            all_potential_sources.sort(key=lambda x: x.get("presentation_likelihood", 0), reverse=True)
            
            # Create comprehensive statistics
            total_sources = len(all_potential_sources)
            high_likelihood_sources = len([s for s in all_potential_sources if s.get("presentation_likelihood", 0) >= 0.7])
            medium_likelihood_sources = len([s for s in all_potential_sources if 0.4 <= s.get("presentation_likelihood", 0) < 0.7])
            
            # Group by data categories
            category_breakdown = {}
            for source in all_potential_sources:
                category = source.get("value_category", "unknown")
                category_breakdown[category] = category_breakdown.get(category, 0) + 1
            
            workbook_summary = {
                "workbook_summary": {
                    "total_sheets_processed": len(sheet_analyses),
                    "total_potential_sources": total_sources,
                    "high_likelihood_sources": high_likelihood_sources,
                    "medium_likelihood_sources": medium_likelihood_sources,
                    "category_breakdown": category_breakdown,
                    "analysis_timestamp": datetime.utcnow().isoformat(),
                    "comprehensive_extraction": True,
                    "coverage": "FULL - No artificial limits applied"
                },
                "potential_sources": all_potential_sources,  # ALL SOURCES - NO LIMITS
                "sheet_analyses": sheet_analyses,
                "extraction_metadata": {
                    "approach": "comprehensive_batched_analysis",
                    "ai_model": "gemini-2.5-pro",
                    "limitations_removed": [
                        "sheet_count_limit",
                        "cell_range_limits", 
                        "numeric_cell_limits",
                        "final_results_limits"
                    ]
                }
            }
            
            logger.info(f"COMPREHENSIVE workbook synthesis completed:")
            logger.info(f"  - Total sources: {total_sources}")
            logger.info(f"  - High likelihood: {high_likelihood_sources}")
            logger.info(f"  - Medium likelihood: {medium_likelihood_sources}")
            logger.info(f"  - Categories: {list(category_breakdown.keys())}")
            
            return workbook_summary
            
        except Exception as e:
            logger.error(f"Comprehensive workbook synthesis failed: {e}")
            # Return basic structure with all sources
            return {
                "workbook_summary": {
                    "total_sheets_processed": len(sheet_analyses),
                    "total_potential_sources": len(all_potential_sources),
                    "analysis_timestamp": datetime.utcnow().isoformat(),
                    "comprehensive_extraction": True
                },
                "potential_sources": all_potential_sources,  # Still return ALL sources even on error
                "sheet_analyses": sheet_analyses,
                "synthesis_error": str(e)
            }

    # ... [Keep all the other existing methods like run_direct_comprehensive_audit, etc. unchanged] ...
    # [The rest of the methods remain the same as in your original code]

    async def run_direct_comprehensive_audit(self, pdf_values: List[Dict], excel_values: List[Dict]) -> Dict[str, Any]:
        """Run direct comprehensive audit comparing ALL PDF values against ALL Excel values"""
        logger.info(f"Starting direct comprehensive audit: {len(pdf_values)} PDF values vs {len(excel_values)} Excel values")
        
        if not pdf_values or not excel_values:
            return {
                "summary": {"total_values_checked": 0, "overall_accuracy": 0.0},
                "detailed_results": [],
                "recommendations": ["No values available for audit"],
                "risk_assessment": "high"
            }
        
        # Process PDF values in smaller batches for better reliability
        batch_size = 5  # Reduced batch size for better JSON reliability
        all_audit_results = []
        
        for i in range(0, len(pdf_values), batch_size):
            batch = pdf_values[i:i+batch_size]
            batch_results = await self._process_direct_audit_batch(batch, excel_values, i // batch_size + 1)
            all_audit_results.extend(batch_results)
            await asyncio.sleep(1)  # Rate limiting
        
        # Calculate comprehensive summary
        summary = {
            "total_values_checked": len(all_audit_results),
            "matched": len([r for r in all_audit_results if r.get("validation_status") == "matched"]),
            "mismatched": len([r for r in all_audit_results if r.get("validation_status") == "mismatched"]),
            "formatting_differences": len([r for r in all_audit_results if r.get("validation_status") == "formatting_difference"]),
            "unverifiable": len([r for r in all_audit_results if r.get("validation_status") == "unverifiable"]),
            "pdf_only": len([r for r in all_audit_results if r.get("validation_status") == "pdf_only"]),
        }
        
        if summary["total_values_checked"] > 0:
            summary["overall_accuracy"] = ((summary["matched"] + summary["formatting_differences"]) / summary["total_values_checked"]) * 100
        else:
            summary["overall_accuracy"] = 0.0

        # Generate comprehensive recommendations
        recommendations = self._generate_direct_audit_recommendations(summary)
        
        return {
            "summary": summary,
            "detailed_results": all_audit_results,
            "recommendations": recommendations,
            "risk_assessment": "low" if summary["overall_accuracy"] >= 85 else "medium" if summary["overall_accuracy"] >= 70 else "high",
            "coverage_analysis": {
                "pdf_values_analyzed": len(pdf_values),
                "excel_values_searched": len(excel_values),
                "coverage_percentage": 100.0,
                "approach": "comprehensive_direct_validation"
            }
        }

    async def _process_direct_audit_batch(self, pdf_batch: List[Dict], all_excel_values: List[Dict], batch_num: int) -> List[Dict]:
        """Process a batch of PDF values against all Excel values"""
        
        # Use MORE Excel values for comprehensive comparison (remove the 30 limit)
        excel_sample = all_excel_values[:100] if len(all_excel_values) > 100 else all_excel_values
        
        prompt = f"""
You are auditing presentation values against Excel source data.

PDF VALUES TO VALIDATE (Batch {batch_num}):
{json.dumps(pdf_batch, indent=1)}

EXCEL VALUES TO SEARCH AGAINST:
{json.dumps(excel_sample, indent=1)}

For each PDF value, find its best match in Excel values and determine validation status.

Return ONLY valid JSON:
{{
    "batch_results": [
        {{
            "pdf_value_id": "pdf_value_id_from_input",
            "pdf_value": "actual_pdf_value",
            "pdf_context": "business_context_meaning",
            "validation_status": "matched|mismatched|formatting_difference|unverifiable|pdf_only",
            "excel_match": {{
                "source_file": "matched_excel_file_name",
                "cell_reference": "matched_cell_reference", 
                "excel_value": "matched_excel_value",
                "match_confidence": 0.95
            }},
            "confidence": 0.95,
            "audit_reasoning": "detailed_explanation_of_validation"
        }}
    ]
}}

Validation Status Guide:
- matched: Values are identical or equivalent
- formatting_difference: Same value, different format (e.g., 1000000 vs 1,000,000)
- mismatched: Values are different but related context
- unverifiable: Cannot determine relationship
- pdf_only: No corresponding Excel value found

Return ONLY valid JSON.
"""

        try:
            response = self.model.generate_content(prompt)
            result = await self._parse_gemini_json_response_robust(response.text, f"direct_audit_batch_{batch_num}")
            
            batch_results = result.get('batch_results', [])
            
            # Combine with original PDF values to ensure completeness
            final_results = []
            for i, pdf_value in enumerate(pdf_batch):
                if i < len(batch_results):
                    audit_result = batch_results[i]
                    # Ensure we have all required fields
                    audit_result.update({
                        "original_pdf_data": pdf_value,
                        "audit_timestamp": datetime.utcnow().isoformat(),
                        "batch_number": batch_num
                    })
                    final_results.append(audit_result)
                else:
                    # Fallback for missing results
                    final_results.append({
                        "pdf_value_id": pdf_value.get("id", f"pdf_value_{i}"),
                        "pdf_value": pdf_value.get("value", "unknown"),
                        "pdf_context": pdf_value.get("business_context", {}).get("semantic_meaning", "unknown"),
                        "validation_status": "unverifiable",
                        "excel_match": None,
                        "confidence": 0.0,
                        "audit_reasoning": "Batch processing incomplete",
                        "original_pdf_data": pdf_value,
                        "audit_timestamp": datetime.utcnow().isoformat(),
                        "batch_number": batch_num
                    })
            
            logger.info(f"Direct audit batch {batch_num}: Processed {len(final_results)} PDF values")
            return final_results
            
        except Exception as e:
            logger.error(f"Direct audit batch {batch_num} failed: {e}")
            # Return fallback results
            fallback_results = []
            for i, pdf_value in enumerate(pdf_batch):
                fallback_results.append({
                    "pdf_value_id": pdf_value.get("id", f"pdf_value_{i}"),
                    "pdf_value": pdf_value.get("value", "unknown"),
                    "pdf_context": pdf_value.get("business_context", {}).get("semantic_meaning", "unknown"),
                    "validation_status": "unverifiable",
                    "excel_match": None,
                    "confidence": 0.0,
                    "audit_reasoning": f"Audit error: {str(e)[:100]}",
                    "original_pdf_data": pdf_value,
                    "audit_timestamp": datetime.utcnow().isoformat(),
                    "batch_number": batch_num,
                    "error": str(e)
                })
            return fallback_results

    def _generate_direct_audit_recommendations(self, summary: Dict) -> List[str]:
        """Generate recommendations based on direct audit summary"""
        recommendations = []
        
        total = summary["total_values_checked"]
        if total == 0:
            return ["No values were audited. Check extraction process."]
        
        accuracy = summary["overall_accuracy"]
        matched = summary["matched"]
        formatting_diffs = summary["formatting_differences"]
        mismatched = summary["mismatched"]
        pdf_only = summary["pdf_only"]
        
        # Overall assessment
        if accuracy >= 90:
            recommendations.append(f"Excellent data quality: {accuracy:.1f}% accuracy across all {total} values.")
        elif accuracy >= 75:
            recommendations.append(f"Good data quality: {accuracy:.1f}% accuracy. Review flagged discrepancies.")
        else:
            recommendations.append(f"Data quality concerns: {accuracy:.1f}% accuracy. Comprehensive review needed.")
        
        # Specific recommendations
        if matched > 0:
            recommendations.append(f"✅ {matched} values perfectly matched between presentation and Excel.")
        
        if formatting_diffs > 0:
            recommendations.append(f"📋 {formatting_diffs} values have formatting differences but same underlying data.")
        
        if mismatched > 0:
            recommendations.append(f"⚠️ {mismatched} values show potential discrepancies requiring review.")
        
        if pdf_only > 0:
            percentage = (pdf_only / total) * 100
            if percentage > 20:
                recommendations.append(f"🔍 {pdf_only} values ({percentage:.1f}%) appear only in presentation - verify if these should have Excel sources.")
            else:
                recommendations.append(f"ℹ️ {pdf_only} values are presentation-only (calculated metrics, external data, etc.)")
        
        # Coverage assessment
        recommendations.append(f"📊 Comprehensive coverage: All {total} extracted values were validated (100% coverage).")
        
        return recommendations

    # ... [Include all other existing methods unchanged] ...

    async def _synthesize_document_analysis(self, page_analyses: List[Dict]) -> Dict[str, Any]:
        """
        Synthesize comprehensive document analysis using Gemini 2.5 Pro
        """
        # Combine all extracted values
        all_values = []
        for page in page_analyses:
            page_values = page.get('extracted_values', [])
            for value in page_values:
                value['page_number'] = page.get('page_number', 0)
                all_values.append(value)

        # Simplified synthesis for better reliability
        try:
            prompt = f"""
Analyze {len(page_analyses)} presentation pages with {len(all_values)} extracted values.

Create a document summary in JSON format:
{{
    "document_summary": {{
        "total_pages": {len(page_analyses)},
        "document_type": "financial_presentation",
        "main_business_themes": ["revenue", "growth", "performance"]
    }},
    "all_extracted_values": {json.dumps(all_values[:100])},
    "extraction_quality_metrics": {{
        "total_values_extracted": {len(all_values)},
        "overall_confidence": 0.85
    }}
}}

Return only valid JSON.
"""

            response = self.model.generate_content(prompt)
            synthesis = await self._parse_gemini_json_response_robust(response.text, "document_synthesis")
            
            # Ensure all_extracted_values is populated
            if not synthesis.get('all_extracted_values'):
                synthesis['all_extracted_values'] = all_values
            
            logger.info("Document synthesis completed successfully")
            return synthesis
            
        except Exception as e:
            logger.error(f"Document synthesis failed: {e}")
            # Return fallback structure
            return {
                "document_summary": {
                    "total_pages": len(page_analyses),
                    "document_type": "financial_presentation"
                },
                "all_extracted_values": all_values,
                "extraction_quality_metrics": {
                    "total_values_extracted": len(all_values),
                    "overall_confidence": 0.8
                },
                "synthesis_error": str(e)
            }

    def _validate_and_enhance_coordinates(self, result: Dict, image_size: Tuple[int, int]) -> Dict:
        """Validate and enhance coordinate data"""
        width, height = image_size
        
        for value in result.get('extracted_values', []):
            if 'coordinates' in value and 'bounding_box' in value['coordinates']:
                bbox = value['coordinates']['bounding_box']
                
                # Ensure coordinates are valid
                if len(bbox) == 4:
                    x1, y1, x2, y2 = bbox
                    x1 = max(0, min(1, float(x1)))
                    y1 = max(0, min(1, float(y1)))
                    x2 = max(0, min(1, float(x2)))
                    y2 = max(0, min(1, float(y2)))
                    
                    # Ensure logical ordering
                    if x1 > x2:
                        x1, x2 = x2, x1
                    if y1 > y2:
                        y1, y2 = y2, y1
                    
                    # Ensure minimum size
                    if x2 - x1 < 0.01:
                        x2 = min(1, x1 + 0.01)
                    if y2 - y1 < 0.01:
                        y2 = min(1, y1 + 0.01)
                    
                    value['coordinates']['bounding_box'] = [x1, y1, x2, y2]
                    value['coordinates']['center_point'] = [(x1 + x2) / 2, (y1 + y2) / 2]
                else:
                    # Set default coordinates if invalid
                    value['coordinates']['bounding_box'] = [0.1, 0.1, 0.2, 0.2]
                    value['coordinates']['center_point'] = [0.15, 0.15]
        
        return result

    async def _parse_gemini_json_response_robust(self, response_text: str, context: str) -> Dict[str, Any]:
        """Robust JSON parsing for Gemini responses with multiple fallback strategies"""
        try:
            # Step 1: Basic cleaning
            cleaned_text = response_text.strip()
            
            # Remove common markdown formatting
            if cleaned_text.startswith("```json"):
                cleaned_text = cleaned_text[7:]
            elif cleaned_text.startswith("```"):
                cleaned_text = cleaned_text[3:]
            
            if cleaned_text.endswith("```"):
                cleaned_text = cleaned_text[:-3]
            
            # Step 2: Find JSON boundaries more carefully
            json_start = -1
            json_end = -1
            brace_count = 0
            
            for i, char in enumerate(cleaned_text):
                if char == '{':
                    if json_start == -1:
                        json_start = i
                    brace_count += 1
                elif char == '}':
                    brace_count -= 1
                    if brace_count == 0 and json_start != -1:
                        json_end = i
                        break
            
            if json_start == -1 or json_end == -1:
                raise ValueError(f"No complete JSON object found in response for {context}")
            
            json_str = cleaned_text[json_start:json_end + 1]
            
            # Step 3: Advanced JSON cleaning
            json_str = self._clean_json_aggressively(json_str)
            
            # Step 4: Try to parse
            result = json.loads(json_str)
            logger.info(f"Successfully parsed Gemini JSON for {context}")
            return result
            
        except json.JSONDecodeError as e:
            logger.error(f"JSON parsing error for {context}: {e}")
            # Try recovery strategies
            return await self._json_recovery_strategies(response_text, context)
            
        except Exception as e:
            logger.error(f"Unexpected parsing error for {context}: {e}")
            return self._get_fallback_structure(context)

    def _clean_json_aggressively(self, json_str: str) -> str:
        """Aggressively clean JSON string for common Gemini issues"""
        # Remove trailing commas before closing braces/brackets
        json_str = re.sub(r',(\s*[}\]])', r'\1', json_str)
        
        # Fix unquoted keys
        json_str = re.sub(r'([{,]\s*)([a-zA-Z_][a-zA-Z0-9_]*)\s*:', r'\1"\2":', json_str)
        
        # Fix single quotes to double quotes
        json_str = re.sub(r"'([^']*)'", r'"\1"', json_str)
        
        # Remove comments if any
        json_str = re.sub(r'//.*?$', '', json_str, flags=re.MULTILINE)
        
        # Fix common value issues
        json_str = re.sub(r':\s*True\b', ': true', json_str)
        json_str = re.sub(r':\s*False\b', ': false', json_str)
        json_str = re.sub(r':\s*None\b', ': null', json_str)
        
        return json_str

    async def _json_recovery_strategies(self, response_text: str, context: str) -> Dict[str, Any]:
        """Multiple strategies to recover from JSON parsing failures"""
        
        # Strategy 1: Try to extract just the core data
        try:
            # Look for key patterns and extract manually
            if "extracted_values" in response_text or "potential_sources" in response_text or "batch_analysis" in response_text:
                # Try to find array content
                for key in ["extracted_values", "potential_sources", "batch_analysis"]:
                    start_pattern = rf'"{key}"\s*:\s*\['
                    
                    match = re.search(start_pattern, response_text)
                    if match:
                        start_pos = match.end() - 1  # Include the [
                        
                        # Find matching closing bracket
                        bracket_count = 0
                        end_pos = -1
                        for i in range(start_pos, len(response_text)):
                            if response_text[i] == '[':
                                bracket_count += 1
                            elif response_text[i] == ']':
                                bracket_count -= 1
                                if bracket_count == 0:
                                    end_pos = i
                                    break
                        
                        if end_pos != -1:
                            array_content = response_text[start_pos:end_pos + 1]
                            # Try to parse just this array
                            try:
                                extracted_array = json.loads(array_content)
                                return {key: extracted_array}
                            except:
                                continue
            
        except:
            pass
        
        # Strategy 2: Return fallback structure
        return self._get_fallback_structure(context)

    def _get_fallback_structure(self, context: str) -> Dict[str, Any]:
        """Return appropriate fallback structure based on context"""
        if "page_" in context:
            return {
                "page_number": 1,
                "page_dimensions": {"width": 800, "height": 600},
                "extracted_values": [],
                "error": f"JSON parsing failed for {context}"
            }
        elif "excel_batch" in context:
            return {
                "batch_analysis": [],
                "error": f"JSON parsing failed for {context}"
            }
        elif "mapping" in context.lower():
            return {
                "suggested_mappings": [],
                "mapping_quality_assessment": {"overall_coverage": 0},
                "error": f"JSON parsing failed for {context}"
            }
        elif "audit" in context.lower():
            return {
                "batch_results": [],
                "error": f"JSON parsing failed for {context}"
            }
        else:
            return {
                "error": f"JSON parsing failed for {context}"
            }

# Initialize the enhanced service
enhanced_gemini_service = EnhancedGeminiService()